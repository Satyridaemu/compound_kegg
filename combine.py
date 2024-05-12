import openpyxl
filename = "KEGG_data.xlsx"  # 替换成你的 Excel 文件名
sheet_name = "sheet"  # 替换成你的工作表名
row3=1
wb = openpyxl.load_workbook(filename)
sheet = wb[sheet_name]
A_count = 0
for cell in sheet['A']:
    # 如果单元格不是空的，增加计数器
    if cell.value is not None:
        A_count += 1
    # 遍历 A 列和 B 列，并组合它们的值，然后写入 C 列
for row in range(1, A_count + 1):
    a_value = sheet.cell(row=row, column=1).value
    for i in range(1, 603):
        b_value = sheet.cell(row=i, column=2).value
        combined_value = f"{a_value}{b_value}"
        sheet.cell(row=row3, column=3).value = combined_value
        row3=row3+1

    # 保存修改后的 Excel 文件
    wb.save(filename)
print(row3)
# 调用函数
