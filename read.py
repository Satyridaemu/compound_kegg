import requests
import re
from openpyxl import load_workbook

# 发送请求获取网页内容
url = "https://www.kegg.jp/kegg/pathway.html"  # 你要爬取的网页链接
response = requests.get(url)
html_content = response.text

# 使用正则表达式匹配所有以0开头的五位数字
pattern = r'\b0\d{4}\b'
matches = re.findall(pattern, html_content)

# 加载现有的Excel文件
wb = load_workbook("KEGG_data.xlsx")
ws = wb.active

# 将匹配到的数字写入Excel的第一列
for idx, match in enumerate(matches, start=1):
    ws.cell(row=idx, column=2, value=match)

# 保存Excel文件
wb.save("KEGG_data.xlsx")