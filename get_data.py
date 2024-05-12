import json
import time

from Bio.KEGG import REST
from openpyxl import load_workbook
from identifer_in_main import find_and_store_values
da={}
fault=0
success=0
count=1
def get_data(i,list_name):
    global fault,success,count,da
    pathway = REST.kegg_get(list_name[i])
    if pathway==None :
        print(f'第{i+1}组失败\t'f'当前进度{i+1}/{total_items}\t','成功:',success,'\t','失败:',fault)
        fault=fault+1
        return
    res = {list_name[i]:pathway.read().split("\n")}
    up_dict=find_and_store_values(res)
    da.update(up_dict)
    print(f'第{i+1}组成功\t'f'当前进度{i+1}/{total_items}\t','成功:',success,'\t','失败:',fault)
    success=success+1
    if (success%100==0):
        count=count+1
        da={}
    with open(f'compound{count}.json', 'w') as f:
        json.dump(da, f,separators=(",\n", ": "))

wb = load_workbook('KEGG_data.xlsx')
wa=wb.active  #读数据
column_to_read = wa['C']
list = [cell.value for cell in column_to_read]
total_items=(wa.max_row)

for item in range(0,total_items):  #1807
    get_data(item, list)
#print()

